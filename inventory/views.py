from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db.models import Max, IntegerField, Count, Min
from django.db.models.functions import Cast
import json
import csv
import urllib.parse
import os
import uuid
from datetime import timedelta
from itertools import chain
from .models import InventoryItem, StatusHistory, NotificationLog, ItemPhoto, ChangeLog, ScanLog, Tag, PrintJob

LOCATION_CHOICES = [
    'York, PA',
    'Rockville, MD',
    'Cambridge, MD',
]


def scanner_landing(request):
    """Main scanner landing page — supports ?data= (QR payload) and ?id= (barcode ID)"""
    barcode_data = request.GET.get('data', '')
    item_id = request.GET.get('id', '')

    if not barcode_data and not item_id:
        return render(request, 'inventory/scanner_landing.html', {
            'error': 'Invalid scan. No data found.',
            'item': None,
            'item_json': '{}'
        })

    try:
        if item_id:
            # Barcode scan — lookup by database ID
            item = get_object_or_404(InventoryItem, id=int(item_id))
        else:
            # QR code scan — parse payload
            decoded_data = urllib.parse.unquote(barcode_data)
            parts = decoded_data.split(' | ')
            data_dict = {}

            for part in parts:
                if '=' in part:
                    key, value = part.split('=', 1)
                    data_dict[key.strip()] = value.strip()

            manufacturer = data_dict.get('MFR', '')
            pallet_id = data_dict.get('PALLET', '')
            box_id = data_dict.get('BOX', '')

            item = get_object_or_404(
                InventoryItem,
                manufacturer=manufacturer,
                pallet_id=pallet_id,
                box_id=int(box_id)
            )

        # Log this scan
        ScanLog.objects.create(item=item)

        history = item.status_history.all()
        status_labels = dict(item.STATUS_CHOICES)
        photos = item.photos.all()
        scan_count = item.scan_logs.count()
        _annotate_qr_urls([item])

        # Build unified audit trail from status history + change logs
        status_entries = list(item.status_history.all())
        change_entries = list(item.change_logs.all())
        for e in status_entries:
            e.entry_type = 'status'
        for e in change_entries:
            e.entry_type = 'change'
        audit_trail = sorted(
            chain(status_entries, change_entries),
            key=lambda e: e.changed_at,
            reverse=True,
        )

        preset_tags = ['Standard', 'Initiating', 'Target', 'RP12', 'RP48',
                       'RP48-25kW', 'RP48-40kW', 'RP240']

        return render(request, 'inventory/scanner_landing.html', {
            'item': item,
            'history': history,
            'audit_trail': audit_trail,
            'status_labels': status_labels,
            'photos': photos,
            'scan_count': scan_count,
            'location_choices': LOCATION_CHOICES,
            'preset_tags': preset_tags,
            'error': None
        })

    except Exception as e:
        return render(request, 'inventory/scanner_landing.html', {
            'error': f'Error loading item: {str(e)}',
            'item': None,
            'item_json': '{}'
        })


@csrf_exempt
@require_http_methods(["POST"])
def update_status(request):
    """Endpoint to update item status with checkout attribution and audit trail"""
    try:
        item_id = request.POST.get('item_id')
        new_status = request.POST.get('status')
        notes = request.POST.get('notes', '')
        changed_by = request.POST.get('changed_by', '')

        status_mapping = {
            'Checked In': 'checked_in',
            'Checked Out': 'checked_out',
            'Tested': 'tested',
            'Will Be Reused': 'will_be_reused',
            'Recycling': 'recycling',
        }

        status_value = status_mapping.get(new_status)
        if not status_value:
            return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)

        item = get_object_or_404(InventoryItem, id=item_id)
        old_status = item.status
        item.status = status_value

        # Checkout attribution
        if status_value == 'checked_out':
            item.checked_out_by = changed_by
            item.checked_out_at = timezone.now()
        elif old_status == 'checked_out' and status_value != 'checked_out':
            item.checked_out_by = ''
            item.checked_out_at = None

        item.save()

        # Record status change history with audit trail
        StatusHistory.objects.create(
            item=item,
            old_status=old_status,
            new_status=status_value,
            notes=notes,
            changed_by=changed_by,
        )

        # Send notification for checkout or damage
        _send_notification(item, old_status, status_value, changed_by)

        return JsonResponse({
            'success': True,
            'status': new_status,
            'last_updated': item.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def _send_notification(item, old_status, new_status, changed_by):
    """Send webhook notifications for checkout/damage events"""
    webhook_url = os.environ.get('NOTIFICATION_WEBHOOK_URL', '')
    notification_type = None
    message = ''

    if new_status == 'checked_out':
        notification_type = 'checkout'
        message = f'{item.manufacturer} Box #{item.box_id} checked out by {changed_by or "unknown"}'
    elif item.damaged and old_status != new_status:
        notification_type = 'damaged'
        message = f'{item.manufacturer} Box #{item.box_id} (DAMAGED) status changed to {dict(InventoryItem.STATUS_CHOICES).get(new_status, new_status)}'
    elif old_status != new_status:
        notification_type = 'status_change'
        message = f'{item.manufacturer} Box #{item.box_id} status: {dict(InventoryItem.STATUS_CHOICES).get(old_status, old_status)} -> {dict(InventoryItem.STATUS_CHOICES).get(new_status, new_status)}'

    if notification_type:
        NotificationLog.objects.create(
            item=item,
            notification_type=notification_type,
            message=message,
            sent_to=webhook_url or 'logged_only',
        )

        if webhook_url:
            try:
                import requests
                requests.post(webhook_url, json={
                    'type': notification_type,
                    'message': message,
                    'item_id': item.id,
                    'manufacturer': item.manufacturer,
                    'box_id': item.box_id,
                    'pallet_id': item.pallet_id,
                    'changed_by': changed_by,
                }, timeout=5)
            except Exception:
                pass  # Don't fail the status update if notification fails


def item_api(request):
    """API endpoint to get item information"""
    manufacturer = request.GET.get('mfr', '')
    pallet_id = request.GET.get('pallet', '')
    box_id = request.GET.get('box', '')

    if not all([manufacturer, pallet_id, box_id]):
        return JsonResponse({'error': 'Missing parameters'}, status=400)

    try:
        item = get_object_or_404(
            InventoryItem,
            manufacturer=manufacturer,
            pallet_id=pallet_id,
            box_id=int(box_id)
        )

        return JsonResponse({
            'id': item.id,
            'manufacturer': item.manufacturer,
            'pallet_id': item.pallet_id,
            'box_id': item.box_id,
            'content': item.content,
            'damaged': item.get_damaged_display(),
            'location': item.location,
            'description': item.description,
            'status': dict(item.STATUS_CHOICES).get(item.status, 'Unknown'),
            'last_updated': item.updated_at.isoformat(),
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def dashboard(request):
    """Dashboard view with full inventory listing"""
    show_archived = request.GET.get('archived', '') == '1'

    if show_archived:
        items = InventoryItem.objects.filter(archived=True).annotate(
            pallet_num=Cast('pallet_id', IntegerField())
        ).order_by('pallet_num', 'manufacturer', 'box_id')
    else:
        items = InventoryItem.objects.filter(archived=False).annotate(
            pallet_num=Cast('pallet_id', IntegerField())
        ).order_by('pallet_num', 'manufacturer', 'box_id')

    all_active = InventoryItem.objects.filter(archived=False)
    cutoff = timezone.now() - timedelta(days=7)
    overdue_qs = all_active.filter(
        status='checked_out',
        checked_out_at__isnull=False,
        checked_out_at__lt=cutoff,
    )
    overdue_count = overdue_qs.count()

    _annotate_qr_urls(items)

    # #19: Activity feed — recent changes across all items
    recent_status = list(StatusHistory.objects.select_related('item').order_by('-changed_at')[:15])
    recent_changes = list(ChangeLog.objects.select_related('item').order_by('-changed_at')[:15])
    for e in recent_status:
        e.entry_type = 'status'
    for e in recent_changes:
        e.entry_type = 'change'
    recent_activity = sorted(
        chain(recent_status, recent_changes),
        key=lambda e: e.changed_at,
        reverse=True,
    )[:20]

    return render(request, 'inventory/dashboard.html', {
        'items': items,
        'checked_in_count': all_active.filter(status='checked_in').count(),
        'checked_out_count': all_active.filter(status='checked_out').count(),
        'damaged_count': all_active.filter(damaged=True).count(),
        'archived_count': InventoryItem.objects.filter(archived=True).count(),
        'overdue_count': overdue_count,
        'overdue_items': overdue_qs,
        'show_archived': show_archived,
        'recent_activity': recent_activity,
        'location_choices': LOCATION_CHOICES,
    })


def item_history(request, item_id):
    """Item history view"""
    item = get_object_or_404(InventoryItem, id=item_id)
    history = item.status_history.all()
    return render(request, 'inventory/item_history.html', {'item': item, 'history': history})


@csrf_exempt
@require_http_methods(["POST"])
def update_history_notes(request):
    """Update notes on a status history entry"""
    try:
        history_id = request.POST.get('history_id')
        notes = request.POST.get('notes', '')

        entry = get_object_or_404(StatusHistory, id=history_id)
        entry.notes = notes
        entry.save()

        return JsonResponse({'success': True, 'notes': entry.notes})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def bulk_update_status(request):
    """Bulk update status for multiple items"""
    try:
        data = json.loads(request.body)
        item_ids = data.get('item_ids', [])
        new_status = data.get('status', '')
        notes = data.get('notes', '')
        changed_by = data.get('changed_by', '')

        if not item_ids or not new_status:
            return JsonResponse({'success': False, 'error': 'Missing item_ids or status'}, status=400)

        valid_statuses = dict(InventoryItem.STATUS_CHOICES)
        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)

        updated = 0
        for item in InventoryItem.objects.filter(id__in=item_ids):
            old_status = item.status
            item.status = new_status

            if new_status == 'checked_out':
                item.checked_out_by = changed_by
                item.checked_out_at = timezone.now()
            elif old_status == 'checked_out':
                item.checked_out_by = ''
                item.checked_out_at = None

            item.save()

            StatusHistory.objects.create(
                item=item,
                old_status=old_status,
                new_status=new_status,
                notes=notes or 'Bulk status update',
                changed_by=changed_by,
            )
            _send_notification(item, old_status, new_status, changed_by)
            updated += 1

        return JsonResponse({'success': True, 'updated_count': updated})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def export_csv(request):
    """Export inventory to CSV"""
    show_archived = request.GET.get('archived', '') == '1'

    if show_archived:
        items = InventoryItem.objects.filter(archived=True).order_by('-updated_at')
    else:
        items = InventoryItem.objects.filter(archived=False).order_by('-updated_at')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_export.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Tag ID', 'Manufacturer', 'Pallet ID', 'Box ID', 'Contents', 'Damaged',
        'Location', 'Description', 'Tags', 'Status', 'Checked Out By', 'Checked Out At',
        'Created', 'Updated', 'Archived', 'Barcode Payload'
    ])

    status_labels = dict(InventoryItem.STATUS_CHOICES)

    for item in items:
        writer.writerow([
            item.id,
            item.tag_id,
            item.manufacturer,
            item.pallet_id,
            item.box_id,
            item.content,
            'Yes' if item.damaged else 'No',
            item.location,
            item.description,
            item.tags,
            status_labels.get(item.status, item.status),
            item.checked_out_by,
            item.checked_out_at.strftime('%Y-%m-%d %H:%M') if item.checked_out_at else '',
            item.created_at.strftime('%Y-%m-%d %H:%M'),
            item.updated_at.strftime('%Y-%m-%d %H:%M'),
            'Yes' if item.archived else 'No',
            item.barcode_payload,
        ])

    return response


def _get_scan_url(item_id, base_url=None):
    """Return the scan URL that the QR code should encode."""
    if base_url is None:
        base_url = os.environ.get("SITE_URL", "https://web-production-57c20.up.railway.app")
    return f"{base_url}/scan/?id={item_id}"


def _get_short_qr_url(item_id, base_url=None):
    """Return the QR code image URL — now served locally instead of external API."""
    return f"/qr/{item_id}/code.png"


def _annotate_qr_urls(items, base_url=None):
    """Add short_qr_url attribute to each item for template use."""
    for item in items:
        item.short_qr_url = _get_short_qr_url(item.id)
    return items


def _generate_qr_bytes(item_id, size=300):
    """Generate QR code image bytes locally using qrcode library."""
    import qrcode
    from io import BytesIO

    scan_url = _get_scan_url(item_id)
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(scan_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


def generate_qr_image(request, item_id):
    """Serve a locally-generated QR code image for an item."""
    get_object_or_404(InventoryItem, id=item_id)
    buf = _generate_qr_bytes(item_id)
    response = HttpResponse(buf.read(), content_type='image/png')
    response['Cache-Control'] = 'public, max-age=86400'
    return response


def _make_labeled_qr_image(item):
    """Helper: landscape QR label — QR on left, text on right. Returns BytesIO PNG or None."""
    from PIL import Image as PilImage, ImageDraw, ImageFont
    from io import BytesIO

    try:
        qr_buf = _generate_qr_bytes(item.id)
        qr_img = PilImage.open(qr_buf).convert('RGB')
    except Exception:
        return None

    # Scale QR to fit label (140px square)
    qr_size = 140
    qr_img = qr_img.resize((qr_size, qr_size), PilImage.LANCZOS)

    padding = 10
    text_area_width = 220
    total_w = padding + qr_size + padding + text_area_width + padding
    total_h = qr_size + padding * 2

    canvas = PilImage.new('RGB', (total_w, total_h), 'white')
    canvas.paste(qr_img, (padding, padding))

    draw = ImageDraw.Draw(canvas)

    try:
        font_large = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
        font_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 13)
    except (OSError, IOError):
        font_large = ImageFont.load_default()
        font_small = font_large

    text_x = padding + qr_size + padding
    text_y = padding + 10

    draw.text((text_x, text_y), item.manufacturer, fill='#333', font=font_large)
    text_y += 28
    draw.text((text_x, text_y), f"Box #{item.box_id}", fill='#333', font=font_small)
    text_y += 22
    draw.text((text_x, text_y), f"Pallet {item.pallet_id}", fill='#333', font=font_small)
    text_y += 22
    project = getattr(item, 'project_number', '') or ''
    if project:
        draw.text((text_x, text_y), f"Project {project}", fill='#333', font=font_small)

    buf = BytesIO()
    canvas.save(buf, format='PNG')
    buf.seek(0)
    return buf


def export_qr_codes(request):
    """Download all inventory QR codes as an Excel file with embedded QR images"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from io import BytesIO
    import requests as http_requests

    show_archived = request.GET.get('archived', '') == '1'

    if show_archived:
        items = InventoryItem.objects.filter(archived=True).annotate(
            pallet_num=Cast('pallet_id', IntegerField())
        ).order_by('pallet_num', 'manufacturer', 'box_id')
    else:
        items = InventoryItem.objects.filter(archived=False).annotate(
            pallet_num=Cast('pallet_id', IntegerField())
        ).order_by('pallet_num', 'manufacturer', 'box_id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QR Codes'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='8B1A1A', end_color='8B1A1A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = ['Manufacturer', 'Pallet ID', 'Box ID', 'Contents (Qty)', 'Damaged',
               'Location', 'Status', 'QR Code', 'Label']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    base_url = os.environ.get("SITE_URL", "https://web-production-57c20.up.railway.app")
    status_labels = dict(InventoryItem.STATUS_CHOICES)

    for row_num, item in enumerate(items, 2):
        row_data = [
            item.manufacturer,
            item.pallet_id,
            item.box_id,
            item.content,
            'Yes' if item.damaged else 'No',
            item.location,
            status_labels.get(item.status, item.status),
            '',  # QR Code column - image will be inserted
            f"{item.manufacturer}\nPallet {item.pallet_id}\nBox #{item.box_id}",
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num == 9:  # Label column
                cell.alignment = center_alignment

        # Set row height for labeled QR code image
        ws.row_dimensions[row_num].height = 95

        # Generate and embed labeled QR code image
        if item.qr_url:
            labeled_buf = _make_labeled_qr_image(item)
            if labeled_buf:
                img = XlImage(labeled_buf)
                img.width = 220
                img.height = 95
                cell_ref = f'H{row_num}'
                ws.add_image(img, cell_ref)
            else:
                ws.cell(row=row_num, column=8, value=item.qr_url)

    # Set column widths
    col_widths = {'A': 18, 'B': 10, 'C': 8, 'D': 14, 'E': 10,
                  'F': 16, 'G': 14, 'H': 30, 'I': 22}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    label = 'archived' if show_archived else 'inventory'
    response['Content-Disposition'] = f'attachment; filename="qr_codes_{label}.xlsx"'
    wb.save(response)
    return response


def export_pdf(request):
    """Export inventory to a simple HTML-based printable report (PDF-ready)"""
    show_archived = request.GET.get('archived', '') == '1'

    if show_archived:
        items = InventoryItem.objects.filter(archived=True).order_by('-updated_at')
    else:
        items = InventoryItem.objects.filter(archived=False).order_by('-updated_at')

    return render(request, 'inventory/export_pdf.html', {
        'items': items,
        'generated_at': timezone.now(),
        'status_labels': dict(InventoryItem.STATUS_CHOICES),
    })


@csrf_exempt
@require_http_methods(["POST"])
def archive_item(request):
    """Archive or unarchive an item"""
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        archive = data.get('archive', True)

        item = get_object_or_404(InventoryItem, id=item_id)
        was_archived = item.archived
        if was_archived == archive:
            return JsonResponse({'success': True, 'archived': item.archived})

        item.archived = archive
        item.archived_at = timezone.now() if archive else None
        item.save()

        ChangeLog.objects.create(
            item=item,
            change_type='field_edit',
            field_name='archived',
            old_value='Yes' if was_archived else 'No',
            new_value='Yes' if archive else 'No',
        )

        return JsonResponse({'success': True, 'archived': item.archived})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def bulk_archive(request):
    """Archive multiple items at once"""
    try:
        data = json.loads(request.body)
        item_ids = data.get('item_ids', [])
        archive = data.get('archive', True)

        now = timezone.now() if archive else None
        # Only update items that actually need changing
        items = InventoryItem.objects.filter(id__in=item_ids).exclude(archived=archive)
        changed_ids = list(items.values_list('id', flat=True))
        items.update(archived=archive, archived_at=now)

        for item in InventoryItem.objects.filter(id__in=changed_ids):
            ChangeLog.objects.create(
                item=item,
                change_type='field_edit',
                field_name='archived',
                old_value='No' if archive else 'Yes',
                new_value='Yes' if archive else 'No',
            )

        return JsonResponse({'success': True, 'updated_count': len(changed_ids)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def overdue_items_api(request):
    """API endpoint returning overdue checked-out items"""
    cutoff = timezone.now() - timedelta(days=7)
    items = InventoryItem.objects.filter(
        status='checked_out',
        checked_out_at__isnull=False,
        checked_out_at__lt=cutoff,
        archived=False,
    )

    overdue = []
    for item in items:
        overdue.append({
            'id': item.id,
            'manufacturer': item.manufacturer,
            'pallet_id': item.pallet_id,
            'box_id': item.box_id,
            'checked_out_by': item.checked_out_by,
            'checked_out_at': item.checked_out_at.isoformat(),
            'days_out': item.days_checked_out,
        })

    return JsonResponse({'overdue_items': overdue, 'count': len(overdue)})


def inventory_report_api(request):
    """API endpoint for scheduled inventory report data"""
    active = InventoryItem.objects.filter(archived=False)

    status_counts = {}
    for key, label in InventoryItem.STATUS_CHOICES:
        status_counts[label] = active.filter(status=key).count()

    cutoff = timezone.now() - timedelta(days=7)
    overdue_count = active.filter(
        status='checked_out',
        checked_out_at__isnull=False,
        checked_out_at__lt=cutoff,
    ).count()

    return JsonResponse({
        'total_items': active.count(),
        'status_breakdown': status_counts,
        'damaged_count': active.filter(damaged=True).count(),
        'overdue_count': overdue_count,
        'archived_count': InventoryItem.objects.filter(archived=True).count(),
        'generated_at': timezone.now().isoformat(),
    })


# ---- Shipment Form Views ----

def _next_pallet_id(manufacturer=None):
    """Compute the next pallet ID (max existing + 1). Global across all manufacturers."""
    result = InventoryItem.objects.aggregate(max_pallet=Max(Cast('pallet_id', IntegerField())))
    max_pallet = result['max_pallet']
    if max_pallet:
        try:
            return str(int(max_pallet) + 1)
        except (ValueError, TypeError):
            pass
    return '1'


def add_shipment(request):
    """Single-page form to create a new shipment (replaces Microsoft Form + Excel script)"""
    if request.method == 'GET':
        next_pallet = _next_pallet_id()
        return render(request, 'inventory/add_shipment.html', {
            'next_pallet_id': next_pallet,
            'location_choices': LOCATION_CHOICES,
        })

    # POST — process the form
    manufacturer = request.POST.get('manufacturer', '').strip()
    project_number = request.POST.get('project_number', '').strip()
    # Auto-increment pallet ID
    pallet_id = _next_pallet_id()
    num_boxes = request.POST.get('num_boxes', '').strip()
    items_per_box = request.POST.get('items_per_box', '').strip()
    # Location: use custom if "other" is selected
    location = request.POST.get('location', '').strip()
    if location == 'other':
        location = request.POST.get('location_custom', '').strip()
    damaged = request.POST.get('damaged', 'no')
    description = request.POST.get('description', '').strip()
    damaged_boxes_str = request.POST.get('damaged_boxes', '').strip()
    tags = request.POST.get('tags', '').strip()

    form_data = {
        'manufacturer': manufacturer,
        'pallet_id': pallet_id,
        'project_number': project_number,
        'num_boxes': num_boxes,
        'items_per_box': items_per_box,
        'location': location,
        'damaged': damaged,
        'description': description,
        'damaged_boxes': damaged_boxes_str,
        'tags': tags,
    }

    errors = []

    # Validate required fields
    if not manufacturer:
        errors.append('Manufacturer / Supplier is required.')
    if not location:
        errors.append('Receiving Location is required.')

    try:
        num_boxes_int = int(num_boxes)
        if num_boxes_int < 1:
            errors.append('Number of boxes must be at least 1.')
    except (ValueError, TypeError):
        errors.append('Number of boxes must be a valid number.')
        num_boxes_int = 0

    try:
        items_per_box_int = int(items_per_box)
        if items_per_box_int < 1:
            errors.append('Items per box must be at least 1.')
    except (ValueError, TypeError):
        errors.append('Items per box must be a valid number.')
        items_per_box_int = 0

    # Parse damaged box numbers
    damaged_box_set = set()
    if damaged_boxes_str:
        for part in damaged_boxes_str.split(','):
            part = part.strip()
            if part:
                try:
                    box_num = int(part)
                    if num_boxes_int and (box_num < 1 or box_num > num_boxes_int):
                        errors.append(f'Damaged box #{box_num} is outside the range 1-{num_boxes_int}.')
                    else:
                        damaged_box_set.add(box_num)
                except ValueError:
                    errors.append(f'Invalid damaged box number: "{part}".')

    if errors:
        return render(request, 'inventory/add_shipment.html', {
            'errors': errors,
            'form_data': form_data,
            'next_pallet_id': pallet_id,
            'location_choices': LOCATION_CHOICES,
        })

    # Create items
    base_url = os.environ.get("SITE_URL", "https://web-production-57c20.up.railway.app")
    created_items = []
    shipment_key = str(uuid.uuid4())[:8]

    for box_num in range(1, num_boxes_int + 1):
        box_content = items_per_box_int
        # If specific damaged boxes were listed, only those are damaged.
        # Otherwise fall back to the global "Damage Reported?" flag.
        if damaged_box_set:
            box_damaged = box_num in damaged_box_set
        else:
            box_damaged = damaged == 'yes'

        barcode_payload = f"MFR={manufacturer} | PALLET={pallet_id} | BOX={box_num}"
        encoded_payload = urllib.parse.quote(barcode_payload)
        scanner_url = f"{base_url}/scan/?data={encoded_payload}"
        qr_url = ""  # Will be set after item creation with local endpoint

        # Check if item already exists (update it if so)
        existing = InventoryItem.objects.filter(
            manufacturer=manufacturer,
            pallet_id=pallet_id,
            box_id=box_num
        ).first()

        if existing:
            existing.content = box_content
            existing.damaged = box_damaged
            existing.location = location
            existing.description = description
            existing.project_number = project_number
            existing.tags = tags
            existing.barcode_payload = barcode_payload
            existing.qr_url = _get_short_qr_url(existing.id, base_url)
            existing.save()
            created_items.append(existing)
        else:
            item = InventoryItem.objects.create(
                manufacturer=manufacturer,
                pallet_id=pallet_id,
                box_id=box_num,
                project_number=project_number,
                content=box_content,
                damaged=box_damaged,
                location=location,
                description=description,
                tags=tags,
                status='checked_in',
                barcode_payload=barcode_payload,
                qr_url='',
            )
            # Set short QR URL now that we have the item ID
            item.qr_url = _get_short_qr_url(item.id, base_url)
            item.save()

            # Create initial audit trail entries
            StatusHistory.objects.create(
                item=item,
                old_status='',
                new_status='checked_in',
                notes=f'Item created via shipment (Pallet {pallet_id})',
                changed_by='',
            )
            ChangeLog.objects.create(
                item=item,
                change_type='created',
                field_name='status',
                old_value='',
                new_value='Checked In',
            )

            created_items.append(item)

    # Handle photo uploads — attach to all created items
    photos = request.FILES.getlist('photos')
    for photo_file in photos:
        for item in created_items:
            ItemPhoto.objects.create(
                item=item,
                image=photo_file,
                caption=f'Shipment photo - Pallet {pallet_id}',
            )

    # Store the shipment key in the session for downloads
    request.session[f'shipment_{shipment_key}'] = [item.id for item in created_items]

    _annotate_qr_urls(created_items, base_url)

    return render(request, 'inventory/shipment_result.html', {
        'items': created_items,
        'manufacturer': manufacturer,
        'pallet_id': pallet_id,
        'shipment_key': shipment_key,
    })


def download_shipment_excel(request, shipment_key):
    """Download shipment items as Excel file for QR code printer"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    item_ids = request.session.get(f'shipment_{shipment_key}', [])
    if not item_ids:
        return HttpResponse('Shipment not found or session expired.', status=404)

    items = InventoryItem.objects.filter(id__in=item_ids).order_by('box_id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Shipment Items'

    # Header styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='8B1A1A', end_color='8B1A1A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = ['Box ID', 'Manufacturer', 'Pallet ID', 'Contents (Qty)', 'Damaged',
               'Location', 'Description', 'Barcode Payload', 'Scanner URL', 'QR Code URL']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_num, item in enumerate(items, 2):
        base_url = os.environ.get("SITE_URL", "https://web-production-57c20.up.railway.app")
        encoded_payload = urllib.parse.quote(item.barcode_payload)
        scanner_url = f"{base_url}/scan/?data={encoded_payload}"

        row_data = [
            item.box_id,
            item.manufacturer,
            item.pallet_id,
            item.content,
            'Yes' if item.damaged else 'No',
            item.location,
            item.description,
            item.barcode_payload,
            scanner_url,
            item.qr_url,
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    first_item = items.first()
    if first_item:
        filename = f"shipment_{first_item.manufacturer}_{first_item.pallet_id}.xlsx"
    else:
        filename = "shipment_items.xlsx"
    filename = filename.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@csrf_exempt
@require_http_methods(["POST"])
def edit_item(request):
    """Edit individual item fields (content, damaged, location, description, manufacturer)"""
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        item = get_object_or_404(InventoryItem, id=item_id)
        changed_by = data.get('changed_by', '')

        # Capture old values before any changes
        old_vals = {
            'content': str(item.content),
            'damaged': 'Yes' if item.damaged else 'No',
            'location': item.location,
            'description': item.description,
            'project_number': item.project_number,
            'manufacturer': item.manufacturer,
            'tags': item.tags,
        }

        if 'content' in data:
            try:
                item.content = int(data['content'])
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Contents must be a whole number.'}, status=400)
        if 'damaged' in data:
            item.damaged = data['damaged']
        if 'location' in data:
            item.location = data['location']
        if 'description' in data:
            item.description = data['description']
        if 'project_number' in data:
            item.project_number = data['project_number']
        if 'manufacturer' in data:
            new_mfr = data['manufacturer'].strip()
            if new_mfr:
                item.manufacturer = new_mfr
                # Update barcode payload if manufacturer changed
                if old_vals['manufacturer'] != new_mfr:
                    barcode_payload = f"MFR={new_mfr} | PALLET={item.pallet_id} | BOX={item.box_id}"
                    item.barcode_payload = barcode_payload
                    item.qr_url = _get_short_qr_url(item.id)
        if 'tags' in data:
            item.tags = data['tags']

        # Handle status change within edit (requires Save Changes to log)
        status_changed = False
        old_status = item.status
        if 'status' in data:
            new_status = data['status']
            valid_statuses = dict(InventoryItem.STATUS_CHOICES)
            if new_status in valid_statuses and new_status != old_status:
                item.status = new_status
                status_changed = True
                # Checkout attribution
                if new_status == 'checked_out':
                    item.checked_out_by = changed_by
                    item.checked_out_at = timezone.now()
                elif old_status == 'checked_out':
                    item.checked_out_by = ''
                    item.checked_out_at = None

        item.save()

        # Record status history if status was changed
        if status_changed:
            StatusHistory.objects.create(
                item=item,
                old_status=old_status,
                new_status=item.status,
                notes=data.get('notes', ''),
                changed_by=changed_by,
            )
            _send_notification(item, old_status, item.status, changed_by)

        # Log all field changes to ChangeLog
        new_vals = {
            'content': str(item.content),
            'damaged': 'Yes' if item.damaged else 'No',
            'location': item.location,
            'description': item.description,
            'project_number': item.project_number,
            'manufacturer': item.manufacturer,
            'tags': item.tags,
        }
        for field_name, old_val in old_vals.items():
            new_val = new_vals[field_name]
            if old_val != new_val:
                ChangeLog.objects.create(
                    item=item,
                    change_type='field_edit',
                    field_name=field_name,
                    old_value=old_val,
                    new_value=new_val,
                    changed_by=changed_by,
                )

        return JsonResponse({
            'success': True,
            'manufacturer': item.manufacturer,
            'content': item.content,
            'damaged': item.damaged,
            'location': item.location,
            'description': item.description,
            'project_number': item.project_number,
            'tags': item.tags,
            'status': item.status,
            'status_changed': status_changed,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def upload_photo(request):
    """Upload a photo for an individual item"""
    try:
        item_id = request.POST.get('item_id')
        caption = request.POST.get('caption', '')
        item = get_object_or_404(InventoryItem, id=item_id)

        photo_file = request.FILES.get('photo')
        if not photo_file:
            return JsonResponse({'success': False, 'error': 'No photo file provided'}, status=400)

        # Validate file size (max 10MB)
        if photo_file.size > 10 * 1024 * 1024:
            return JsonResponse({'success': False, 'error': 'Photo must be under 10MB.'}, status=400)

        # Validate image type
        allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        if photo_file.content_type not in allowed_types:
            return JsonResponse({'success': False, 'error': 'Only JPEG, PNG, GIF, and WebP images are allowed.'}, status=400)

        photo = ItemPhoto.objects.create(
            item=item,
            image=photo_file,
            caption=caption,
        )

        return JsonResponse({
            'success': True,
            'photo_id': photo.id,
            'photo_url': photo.image.url,
            'caption': photo.caption,
            'uploaded_at': photo.uploaded_at.strftime('%b %d, %Y %H:%M'),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def delete_photo(request):
    """Delete a photo"""
    try:
        data = json.loads(request.body)
        photo_id = data.get('photo_id')
        photo = get_object_or_404(ItemPhoto, id=photo_id)
        photo.image.delete(save=False)
        photo.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def next_pallet_api(request):
    """API endpoint returning the next pallet ID"""
    next_id = _next_pallet_id()
    return JsonResponse({'next_pallet_id': next_id})


def generate_labeled_qr(request, item_id):
    """Generate a landscape QR label image: QR on left, text on right."""
    from io import BytesIO

    item = get_object_or_404(InventoryItem, id=item_id)

    buf = _make_labeled_qr_image(item)
    if not buf:
        return HttpResponse('Failed to generate QR image', status=500)

    filename = f"QR_{item.manufacturer}_Pallet{item.pallet_id}_Box{item.box_id}.png"
    response = HttpResponse(buf.read(), content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def download_pallet_qr(request, manufacturer, pallet_id):
    """Download QR codes for a specific pallet/shipment as Excel with embedded QR images."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage

    items = InventoryItem.objects.filter(
        manufacturer=manufacturer,
        pallet_id=pallet_id,
    ).order_by('box_id')

    if not items.exists():
        return HttpResponse('No items found for this pallet.', status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QR Codes'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='8B1A1A', end_color='8B1A1A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['Tag ID', 'Manufacturer', 'Pallet ID', 'Box ID', 'Contents (Qty)',
               'Status', 'Tags', 'QR Code', 'Label']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    status_labels = dict(InventoryItem.STATUS_CHOICES)

    for row_num, item in enumerate(items, 2):
        row_data = [
            item.tag_id,
            item.manufacturer,
            item.pallet_id,
            item.box_id,
            item.content,
            status_labels.get(item.status, item.status),
            item.tags,
            '',
            f"{item.manufacturer}\nPallet {item.pallet_id}\nBox #{item.box_id}",
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num == 9:
                cell.alignment = center_alignment

        ws.row_dimensions[row_num].height = 95

        if item.qr_url:
            labeled_buf = _make_labeled_qr_image(item)
            if labeled_buf:
                img = XlImage(labeled_buf)
                img.width = 220
                img.height = 95
                ws.add_image(img, f'H{row_num}')
            else:
                ws.cell(row=row_num, column=8, value=item.qr_url)

    col_widths = {'A': 16, 'B': 18, 'C': 10, 'D': 8, 'E': 14,
                  'F': 14, 'G': 20, 'H': 30, 'I': 22}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_mfr = manufacturer.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="QR_{safe_mfr}_Pallet{pallet_id}.xlsx"'
    wb.save(response)
    return response


def download_shipment_csv(request, shipment_key):
    """Download shipment items as CSV file"""
    item_ids = request.session.get(f'shipment_{shipment_key}', [])
    if not item_ids:
        return HttpResponse('Shipment not found or session expired.', status=404)

    items = InventoryItem.objects.filter(id__in=item_ids).order_by('box_id')

    first_item = items.first()
    if first_item:
        filename = f"shipment_{first_item.manufacturer}_{first_item.pallet_id}.csv"
    else:
        filename = "shipment_items.csv"
    filename = filename.replace(' ', '_')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(['Box ID', 'Manufacturer', 'Pallet ID', 'Contents (Qty)', 'Damaged',
                     'Location', 'Description', 'Barcode Payload', 'Scanner URL', 'QR Code URL'])

    base_url = os.environ.get("SITE_URL", "https://web-production-57c20.up.railway.app")

    for item in items:
        encoded_payload = urllib.parse.quote(item.barcode_payload)
        scanner_url = f"{base_url}/scan/?data={encoded_payload}"
        writer.writerow([
            item.box_id,
            item.manufacturer,
            item.pallet_id,
            item.content,
            'Yes' if item.damaged else 'No',
            item.location,
            item.description,
            item.barcode_payload,
            scanner_url,
            item.qr_url,
        ])

    return response


# ---- Bulk Edit (#7) ----

@csrf_exempt
@require_http_methods(["POST"])
def bulk_edit(request):
    """Bulk edit location, damage, or tags for multiple items."""
    try:
        data = json.loads(request.body)
        item_ids = data.get('item_ids', [])
        fields = data.get('fields', {})
        changed_by = data.get('changed_by', '')

        if not item_ids or not fields:
            return JsonResponse({'success': False, 'error': 'Missing item_ids or fields'}, status=400)

        updated = 0
        for item in InventoryItem.objects.filter(id__in=item_ids):
            if 'location' in fields and fields['location']:
                old_val = item.location
                item.location = fields['location']
                if old_val != item.location:
                    ChangeLog.objects.create(
                        item=item, change_type='field_edit', field_name='location',
                        old_value=old_val, new_value=item.location, changed_by=changed_by,
                    )

            if 'damaged' in fields:
                old_val = 'Yes' if item.damaged else 'No'
                item.damaged = fields['damaged']
                new_val = 'Yes' if item.damaged else 'No'
                if old_val != new_val:
                    ChangeLog.objects.create(
                        item=item, change_type='field_edit', field_name='damaged',
                        old_value=old_val, new_value=new_val, changed_by=changed_by,
                    )

            if 'tags' in fields and fields['tags']:
                old_val = item.tags
                item.tags = fields['tags']
                if old_val != item.tags:
                    ChangeLog.objects.create(
                        item=item, change_type='field_edit', field_name='tags',
                        old_value=old_val, new_value=item.tags, changed_by=changed_by,
                    )

            item.save()
            updated += 1

        return JsonResponse({'success': True, 'updated_count': updated})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ---- Shipment History (#10) ----

def shipment_history(request):
    """Show all past shipments grouped by pallet."""
    shipments = (
        InventoryItem.objects
        .values('manufacturer', 'pallet_id', 'project_number')
        .annotate(
            item_count=Count('id'),
            created=Min('created_at'),
        )
        .order_by('-created')
    )
    return render(request, 'inventory/shipment_history.html', {
        'shipments': shipments,
    })


def shipment_detail(request, manufacturer, pallet_id):
    """Show a past shipment's detail page (same layout as after creation)."""
    items = list(
        InventoryItem.objects.filter(manufacturer=manufacturer, pallet_id=pallet_id)
        .order_by('box_id')
    )
    if not items:
        return redirect('inventory:shipment_history')

    shipment_key = str(uuid.uuid4())[:8]
    request.session[f'shipment_{shipment_key}'] = [item.id for item in items]

    _annotate_qr_urls(items)

    return render(request, 'inventory/shipment_result.html', {
        'items': items,
        'manufacturer': manufacturer,
        'pallet_id': pallet_id,
        'shipment_key': shipment_key,
        'is_history': True,
    })


# ---- Tag Management (#20) ----

def tag_management(request):
    """View and manage all tags across inventory."""
    items = InventoryItem.objects.filter(archived=False).values_list('tags', 'updated_at')
    tag_counts = {}
    tag_last_updated = {}
    for tags_str, updated_at in items:
        if tags_str:
            for tag in tags_str.split(','):
                tag = tag.strip()
                if tag:
                    tag_counts[tag] = tag_counts.get(tag, 0) + 1
                    if tag not in tag_last_updated or (updated_at and updated_at > tag_last_updated[tag]):
                        tag_last_updated[tag] = updated_at
    # Include standalone tags from Tag model (0 items if not used yet)
    for standalone in Tag.objects.all():
        if standalone.name not in tag_counts:
            tag_counts[standalone.name] = 0
    # Build list of dicts with name, count
    tag_list = []
    for name, count in tag_counts.items():
        tag_list.append({
            'name': name,
            'count': count,
        })
    tag_list.sort(key=lambda x: (-x['count'], x['name']))
    # Split into top tags (top 5) and other tags
    top_tags = tag_list[:5]
    other_tags = tag_list[5:]
    return render(request, 'inventory/tag_management.html', {
        'tags': tag_list,
        'top_tags': top_tags,
        'other_tags': other_tags,
    })


@csrf_exempt
@require_http_methods(["POST"])
def rename_tag(request):
    """Rename a tag across all items."""
    try:
        data = json.loads(request.body)
        old_name = data.get('old_name', '').strip()
        new_name = data.get('new_name', '').strip()
        if not old_name or not new_name:
            return JsonResponse({'success': False, 'error': 'Both old and new tag names are required.'}, status=400)

        items = InventoryItem.objects.filter(tags__contains=old_name)
        updated = 0
        for item in items:
            tags = [t.strip() for t in item.tags.split(',') if t.strip()]
            if old_name in tags:
                new_tags = [new_name if t == old_name else t for t in tags]
                item.tags = ','.join(new_tags)
                item.save()
                updated += 1
        return JsonResponse({'success': True, 'updated_count': updated})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def delete_tag(request):
    """Remove a tag from all items."""
    try:
        data = json.loads(request.body)
        tag_name = data.get('tag_name', '').strip()
        if not tag_name:
            return JsonResponse({'success': False, 'error': 'Tag name is required.'}, status=400)

        items = InventoryItem.objects.filter(tags__contains=tag_name)
        updated = 0
        for item in items:
            tags = [t.strip() for t in item.tags.split(',') if t.strip()]
            if tag_name in tags:
                new_tags = [t for t in tags if t != tag_name]
                item.tags = ','.join(new_tags)
                item.save()
                updated += 1
        # Also remove from Tag model if it exists
        Tag.objects.filter(name=tag_name).delete()
        return JsonResponse({'success': True, 'updated_count': updated})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def create_tag(request):
    """Create a new standalone tag."""
    try:
        data = json.loads(request.body)
        tag_name = data.get('tag_name', '').strip()
        if not tag_name:
            return JsonResponse({'success': False, 'error': 'Tag name is required.'}, status=400)

        # Check if tag already exists on any item or in the Tag model
        existing_on_items = InventoryItem.objects.filter(tags__contains=tag_name).exists()
        if existing_on_items or Tag.objects.filter(name=tag_name).exists():
            return JsonResponse({'success': False, 'error': 'A tag with this name already exists.'}, status=400)

        Tag.objects.create(name=tag_name)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ---- Wireless Printing API ----

def _check_print_api_auth(request):
    """Verify Bearer token matches PRINT_API_SECRET. Returns None if OK, or JsonResponse with error."""
    from django.conf import settings
    secret = getattr(settings, 'PRINT_API_SECRET', '')
    if not secret:
        return JsonResponse({'error': 'Print API not configured'}, status=503)
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    if not auth_header.startswith('Bearer ') or auth_header[7:] != secret:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    return None


@csrf_exempt
@require_http_methods(["POST"])
def create_print_jobs(request):
    """Create one PrintJob per item. Called by the web UI 'Send to Printer' button."""
    try:
        data = json.loads(request.body)
        item_ids = data.get('item_ids', [])

        if not item_ids:
            return JsonResponse({'success': False, 'error': 'No item IDs provided'}, status=400)

        items = InventoryItem.objects.filter(id__in=item_ids)
        if not items.exists():
            return JsonResponse({'success': False, 'error': 'No matching items found'}, status=404)

        created = []
        for item in items:
            job = PrintJob.objects.create(item=item, status='pending')
            created.append({'id': job.id, 'item_id': item.id, 'tag_id': item.tag_id})

        return JsonResponse({'success': True, 'jobs_created': len(created), 'jobs': created})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def pending_print_jobs(request):
    """Return pending print jobs for the print worker. Requires Bearer auth."""
    auth_err = _check_print_api_auth(request)
    if auth_err:
        return auth_err

    jobs = PrintJob.objects.filter(status='pending').select_related('item').order_by('created_at')
    base_url = os.environ.get("SITE_URL", "https://web-production-57c20.up.railway.app")

    result = []
    for job in jobs:
        result.append({
            'id': job.id,
            'image_url': f"{base_url}/api/print-jobs/{job.id}/label.png",
        })

    return JsonResponse(result, safe=False)


@csrf_exempt
@require_http_methods(["PATCH"])
def update_print_job_status(request, job_id):
    """Update print job status. Called by print worker. Requires Bearer auth."""
    auth_err = _check_print_api_auth(request)
    if auth_err:
        return auth_err

    try:
        data = json.loads(request.body)
        new_status = data.get('status', '')

        if new_status not in ('printed', 'failed'):
            return JsonResponse({'error': 'Status must be "printed" or "failed"'}, status=400)

        job = get_object_or_404(PrintJob, id=job_id)
        job.status = new_status
        if new_status == 'printed':
            job.printed_at = timezone.now()
        elif new_status == 'failed':
            job.error_message = data.get('error', '')
        job.save()

        return JsonResponse({'success': True, 'id': job.id, 'status': job.status})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def print_job_label_image(request, job_id):
    """Serve the QR label PNG for a print job. Requires Bearer auth."""
    auth_err = _check_print_api_auth(request)
    if auth_err:
        return auth_err

    job = get_object_or_404(PrintJob, id=job_id)
    buf = _make_labeled_qr_image(job.item)
    if not buf:
        return HttpResponse('Failed to generate label image', status=500)

    response = HttpResponse(buf.read(), content_type='image/png')
    response['Cache-Control'] = 'no-cache'
    return response
